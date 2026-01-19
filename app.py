"""
Flask application for MAUDE data processing web interface with authentication.
"""
import os
import json
import csv
from flask import Flask, request, render_template, send_file, jsonify, redirect, url_for, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
import tempfile
import io

from backend.processor import MAUDEProcessor
from backend.auth import AuthManager, User
from backend.imdrf_insights import (
    prepare_data_for_insights,
    analyze_imdrf_insights,
    get_top_manufacturers_for_prefix,
    LEVEL_CONFIG,
    get_imdrf_code_counts_all_levels
)
from backend.imdrf_annex_validator import get_annex_status
from backend.txt_to_csv_converter import TxtToCsvConverter, get_txt_preview
from backend.csv_viewer import LargeCSVViewer, get_csv_page, get_csv_info
from config import (
    GROQ_API_KEY, SECRET_KEY, MAIL_SERVER, MAIL_PORT, MAIL_USE_TLS,
    MAIL_USE_SSL, MAIL_USERNAME, MAIL_PASSWORD, MAIL_DEFAULT_SENDER
)

app = Flask(__name__)
app.config['SECRET_KEY'] = SECRET_KEY
# Note: MAX_CONTENT_LENGTH is set to None to allow streaming uploads of 10GB+ files
# Validation is done at the application level for specific routes
app.config['MAX_CONTENT_LENGTH'] = None
# Use /tmp for Vercel (serverless) or temp directory for local
if os.path.exists('/tmp'):
    app.config['UPLOAD_FOLDER'] = os.path.join('/tmp', 'maude_uploads')
elif os.getenv('TMPDIR'):
    app.config['UPLOAD_FOLDER'] = os.path.join(os.getenv('TMPDIR'), 'maude_uploads')
elif os.getenv('TMP'):
    app.config['UPLOAD_FOLDER'] = os.path.join(os.getenv('TMP'), 'maude_uploads')
else:
    app.config['UPLOAD_FOLDER'] = os.path.join(tempfile.gettempdir(), 'maude_uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Flask-Mail configuration
app.config['MAIL_SERVER'] = MAIL_SERVER
app.config['MAIL_PORT'] = MAIL_PORT
app.config['MAIL_USE_TLS'] = MAIL_USE_TLS
app.config['MAIL_USE_SSL'] = MAIL_USE_SSL
app.config['MAIL_USERNAME'] = MAIL_USERNAME
app.config['MAIL_PASSWORD'] = MAIL_PASSWORD
app.config['MAIL_DEFAULT_SENDER'] = MAIL_DEFAULT_SENDER

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

# Initialize Flask-Mail
mail = Mail(app)

# Initialize Auth Manager
auth_manager = AuthManager(SECRET_KEY)

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}


@login_manager.user_loader
def load_user(user_id):
    """Load user for Flask-Login."""
    return auth_manager.get_user(user_id)


def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/api/imdrf-counts/download-xlsx', methods=['POST'])
@login_required
def api_imdrf_counts_download_xlsx():
    """Upload a cleaned file and download IMDRF code counts for all levels as XLSX."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    file_ext = os.path.splitext(file.filename.lower())[1]
    if file_ext not in {'.csv', '.xlsx', '.xls'}:
        return jsonify({'error': 'Invalid file type. Please upload CSV, XLS, or XLSX file.'}), 400

    temp_path = None

    try:
        filename = secure_filename(file.filename)
        file_id = f"{current_user.id}_{os.urandom(8).hex()}"
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f"counts_{file_id}_{filename}")
        file.save(temp_path)

        counts_by_level = get_imdrf_code_counts_all_levels(temp_path)

        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "IMDRF Code Counts"

        bold_font = Font(bold=True)

        for level in [1, 2, 3]:
            level_label = f"LEVEL-{level} Code"
            ws.append([level_label, ""])
            ws.cell(row=ws.max_row, column=1).font = bold_font

            level_counts = counts_by_level.get(level, {})
            for code in sorted(level_counts.keys()):
                ws.append([code, level_counts.get(code, 0)])

            ws.append(["", ""])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='imdrf_code_counts_all_levels.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


@app.route('/api/imdrf-counts/download-csv', methods=['POST'])
@login_required
def api_imdrf_counts_download_csv():
    """Upload a cleaned file and download IMDRF code counts as CSV (two columns)."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    file_ext = os.path.splitext(file.filename.lower())[1]
    if file_ext not in {'.csv', '.xlsx', '.xls'}:
        return jsonify({'error': 'Invalid file type. Please upload CSV, XLS, or XLSX file.'}), 400

    level = request.form.get('level', 'all')
    if level not in {'all', '1', '2', '3'}:
        return jsonify({'error': 'Invalid level selection.'}), 400

    temp_path = None

    try:
        filename = secure_filename(file.filename)
        file_id = f"{current_user.id}_{os.urandom(8).hex()}"
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f"counts_{file_id}_{filename}")
        file.save(temp_path)

        counts_by_level = get_imdrf_code_counts_all_levels(temp_path)

        rows = []
        if level == 'all':
            for level_num in [1, 2, 3]:
                level_counts = counts_by_level.get(level_num, {})
                for code, count in level_counts.items():
                    rows.append((code, count))
        else:
            level_num = int(level)
            level_counts = counts_by_level.get(level_num, {})
            for code, count in level_counts.items():
                rows.append((code, count))

        rows.sort(key=lambda x: x[0])

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['IMDRF Code', 'Count'])
        for code, count in rows:
            writer.writerow([code, count])

        csv_bytes = io.BytesIO(output.getvalue().encode('utf-8'))
        csv_bytes.seek(0)

        level_label = 'all-levels' if level == 'all' else f"level-{level}"
        download_name = f"imdrf_code_counts_{level_label}.csv"

        return send_file(
            csv_bytes,
            as_attachment=True,
            download_name=download_name,
            mimetype='text/csv'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


def send_password_reset_email(email: str, token: str):
    """Send password reset email."""
    try:
        reset_url = f"{request.host_url}reset-password?token={token}"
        
        msg = Message(
            subject='MAUDE Data Processor - Password Reset',
            recipients=[email],
            html=f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h2 style="color: #2563eb;">Password Reset Request</h2>
                    <p>You requested to reset your password for the MAUDE Data Processor.</p>
                    <p>Click the button below to reset your password:</p>
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{reset_url}" 
                           style="background: #2563eb; color: white; padding: 12px 30px; 
                                  text-decoration: none; border-radius: 6px; display: inline-block;">
                            Reset Password
                        </a>
                    </div>
                    <p style="font-size: 12px; color: #666;">
                        Or copy and paste this link into your browser:<br>
                        <a href="{reset_url}" style="color: #2563eb;">{reset_url}</a>
                    </p>
                    <p style="font-size: 12px; color: #666;">
                        This link will expire in 1 hour. If you didn't request this, please ignore this email.
                    </p>
                </div>
            </body>
            </html>
            """
        )
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Error sending email: {str(e)}")
        return False


# Authentication Routes
@app.route('/login')
def login():
    """Render login page."""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    # Expose test accounts to the login template when enabled
    try:
        from config import TEST_ACCOUNTS_ENABLED, TEST_USER_EMAIL, TEST_USER_PASSWORD
        test_accounts = []
        if TEST_ACCOUNTS_ENABLED:
            test_accounts = [{
                'email': TEST_USER_EMAIL,
                'password': TEST_USER_PASSWORD,
                'label': 'Test Account'
            }]
    except Exception:
        test_accounts = []
    return render_template('login.html', test_accounts=test_accounts)


@app.route('/register')
def register():
    """Render registration page."""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    return render_template('register.html')


@app.route('/forgot-password')
def forgot_password():
    """Render forgot password page."""
    return render_template('forgot_password.html')


@app.route('/reset-password')
def reset_password():
    """Render reset password page."""
    token = request.args.get('token')
    if not token:
        return redirect(url_for('forgot_password'))
    return render_template('reset_password.html', token=token)


@app.route('/logout')
@login_required
def logout():
    """Logout user."""
    logout_user()
    return redirect(url_for('login'))


# API Routes
@app.route('/api/login', methods=['POST'])
def api_login():
    """Handle login API request."""
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    
    if not email or not password:
        return jsonify({'error': 'Email and password are required'}), 400
    
    user = auth_manager.authenticate_user(email, password)
    if user:
        login_user(user, remember=True)
        return jsonify({'success': True, 'message': 'Login successful'}), 200
    else:
        return jsonify({'error': 'Invalid email or password'}), 401


@app.route('/api/register', methods=['POST'])
def api_register():
    """Handle registration API request."""
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    name = data.get('name', '').strip()
    
    if not email or not password:
        return jsonify({'error': 'Email and password are required'}), 400
    
    success, message = auth_manager.register_user(email, password, name)
    if success:
        return jsonify({'success': True, 'message': message}), 200
    else:
        return jsonify({'error': message}), 400


@app.route('/api/forgot-password', methods=['POST'])
def api_forgot_password():
    """Handle forgot password API request."""
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    
    if not email:
        return jsonify({'error': 'Email is required'}), 400
    
    # Always return success to prevent email enumeration
    if auth_manager.user_exists(email):
        token = auth_manager.generate_reset_token(email)
        if token:
            send_password_reset_email(email, token)
    
    return jsonify({
        'success': True,
        'message': 'If an account exists with that email, a password reset link has been sent.'
    }), 200


@app.route('/api/reset-password', methods=['POST'])
def api_reset_password():
    """Handle reset password API request."""
    data = request.get_json()
    token = data.get('token', '')
    password = data.get('password', '')
    
    if not token or not password:
        return jsonify({'error': 'Token and password are required'}), 400
    
    success, message = auth_manager.reset_password(token, password)
    if success:
        return jsonify({'success': True, 'message': message}), 200
    else:
        return jsonify({'error': message}), 400


# Protected Routes
@app.route('/')
@login_required
def index():
    """Render main upload page."""
    return render_template('index.html', user=current_user)


@app.route('/imdrf-code-counts')
@login_required
def imdrf_code_counts_page():
    """Render IMDRF code counts download page."""
    return render_template('imdrf_code_counts.html', user=current_user)


@app.route('/process', methods=['POST'])
@login_required
def process_file():
    """Process uploaded MAUDE file."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Please upload CSV or Excel file.'}), 400
    
    # Note: GROQ_API_KEY is optional - app works without it using deterministic methods only
    # if not GROQ_API_KEY:
    #     return jsonify({'error': 'GROQ_API_KEY not configured. Please set environment variable.'}), 500
    
    try:
        # Save uploaded file
        filename = secure_filename(file.filename)
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(input_path)
        
        # Create output file path
        output_filename = f"cleaned_{filename.rsplit('.', 1)[0]}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        
        # Initialize processor
        processor = MAUDEProcessor()
        
        # Load IMDRF structure if provided
        imdrf_loaded = False
        if 'imdrf_file' in request.files:
            imdrf_file = request.files['imdrf_file']
            if imdrf_file.filename and imdrf_file.filename.strip():
                imdrf_path = os.path.join(app.config['UPLOAD_FOLDER'], f"imdrf_{secure_filename(imdrf_file.filename)}")
                imdrf_file.save(imdrf_path)
                processor.load_imdrf_structure(imdrf_path)
                imdrf_loaded = True
                print(f"IMDRF annex file loaded from: {imdrf_file.filename}")
        
        if not imdrf_loaded:
            print("WARNING: No IMDRF annex file provided. IMDRF codes will be blank.")
        
        # Process file
        stats = processor.process_file(input_path, output_path)
        
        # Check validation - HARD STOPS for critical failures
        critical_failures = []
        if not stats['validation'].get('column_count_correct', False):
            critical_failures.append('Column count check failed')
        if not stats['validation'].get('file_will_open', False):
            critical_failures.append('File integrity check failed')
        if not stats['validation'].get('date_format_correct', False):
            critical_failures.append('Date format check failed (HARD STOP: no literal "nan", all dates must be DD-MM-YYYY)')
        if not stats['validation'].get('imdrf_adjacent', False):
            critical_failures.append('IMDRF Code column position check failed (HARD STOP: must be adjacent to Device Problem)')
        if not stats['validation'].get('imdrf_codes_valid', False):
            critical_failures.append('IMDRF codes validation failed (HARD STOP: all codes must exist in Annex)')
        
        # Non-critical validations (warn but don't fail)
        warnings = []
        if not stats['validation'].get('no_timestamps', False):
            warnings.append('Timestamp check failed (non-critical)')
        
        # HARD STOP on critical failures
        if critical_failures:
            error_msg = f"Validation failed (HARD STOP): {', '.join(critical_failures)}"
            if warnings:
                error_msg += f" | Warnings: {', '.join(warnings)}"
            
            return jsonify({
                'error': error_msg,
                'validation_results': stats['validation'],
                'stats': stats,
                'failed_checks': critical_failures,
                'warnings': warnings
            }), 400
        elif warnings:
            # Log warnings but don't fail
            print(f"Validation warnings: {', '.join(warnings)}")
        
        # Return file for download
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Processing failed: {str(e)}'}), 500


# IMDRF Insights Routes
@app.route('/imdrf-insights')
@login_required
def imdrf_insights_page():
    """Render IMDRF Insights page."""
    return render_template('imdrf_insights.html', user=current_user)


@app.route('/api/imdrf-insights/prepare', methods=['POST'])
@login_required
def api_prepare_insights():
    """Prepare uploaded CSV or Excel file for IMDRF insights analysis."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    # Check file extension
    allowed_extensions = {'.csv', '.xlsx', '.xls'}
    file_ext = os.path.splitext(file.filename.lower())[1]
    if file_ext not in allowed_extensions:
        return jsonify({'error': 'Invalid file type. Please upload CSV, XLS, or XLSX file.'}), 400

    # Get level from form data (default to 1 for backward compatibility)
    level = int(request.form.get('level', 1))
    if level not in [1, 2, 3]:
        return jsonify({'error': 'Invalid level. Must be 1, 2, or 3.'}), 400

    try:
        # Save uploaded file with unique name
        filename = secure_filename(file.filename)
        file_id = f"{current_user.id}_{os.urandom(8).hex()}"
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")
        file.save(input_path)

        # Prepare data for insights at the specified level
        result = prepare_data_for_insights(input_path, level=level)

        # Store file info in session for later use (level can be toggled later)
        session[f'insights_file_{file_id}'] = {
            'path': input_path
        }

        return jsonify({
            'success': True,
            'file_id': file_id,
            'all_prefixes': result['all_prefixes'],
            'all_manufacturers': result['all_manufacturers'],
            'prefix_counts': result.get('prefix_counts', {}),
            'total_rows': result['total_rows'],
            'rows_with_imdrf': result['rows_with_imdrf'],
            'rows_with_dates': result['rows_with_dates'],
            'level': level,
            'level_label': result.get('level_label', f'Level-{level}')
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/refresh', methods=['POST'])
@login_required
def api_refresh_insights():
    """Re-prepare IMDRF insights data for a new level without re-uploading."""
    data = request.get_json() or {}

    file_id = data.get('file_id')
    level = int(data.get('level', 1))

    if not file_id:
        return jsonify({'error': 'Missing file_id parameter'}), 400

    if level not in [1, 2, 3]:
        return jsonify({'error': 'Invalid level. Must be 1, 2, or 3.'}), 400

    try:
        file_info = session.get(f'insights_file_{file_id}')
        if not file_info:
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        if isinstance(file_info, str):
            file_path = file_info
        else:
            file_path = file_info.get('path')

        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        result = prepare_data_for_insights(file_path, level=level)

        return jsonify({
            'success': True,
            'file_id': file_id,
            'all_prefixes': result['all_prefixes'],
            'all_manufacturers': result['all_manufacturers'],
            'prefix_counts': result.get('prefix_counts', {}),
            'total_rows': result['total_rows'],
            'rows_with_imdrf': result['rows_with_imdrf'],
            'rows_with_dates': result['rows_with_dates'],
            'level': level,
            'level_label': result.get('level_label', f'Level-{level}')
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/top-manufacturers', methods=['GET'])
@login_required
def api_top_manufacturers():
    """Get top manufacturers for a specific IMDRF prefix."""
    prefix = request.args.get('prefix')
    file_id = request.args.get('file_id')

    if not prefix or not file_id:
        return jsonify({'error': 'Missing prefix or file_id parameter'}), 400

    try:
        # Retrieve file info from session
        file_info = session.get(f'insights_file_{file_id}')
        if not file_info:
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        # Handle both old format (string) and new format (dict)
        if isinstance(file_info, str):
            file_path = file_info
        else:
            file_path = file_info.get('path')

        level = int(request.args.get('level', 1))
        if level not in [1, 2, 3]:
            return jsonify({'error': 'Invalid level. Must be 1, 2, or 3.'}), 400

        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        # Prepare data again (in memory) at the same level
        result = prepare_data_for_insights(file_path, level=level)
        df_exploded = result['df_exploded']
        mfr_col = result['mfr_col']

        # Get top manufacturers
        top_mfrs = get_top_manufacturers_for_prefix(df_exploded, prefix, mfr_col, top_n=5)

        return jsonify({
            'success': True,
            'top_manufacturers': top_mfrs
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/analyze', methods=['POST'])
@login_required
def api_analyze_insights():
    """Perform IMDRF insights analysis."""
    data = request.get_json()

    file_id = data.get('file_id')
    prefix = data.get('prefix')
    manufacturers = data.get('manufacturers', [])
    grain = data.get('grain', 'W')
    threshold_k = data.get('threshold_k', 2.0)

    if not file_id or not prefix or not manufacturers:
        return jsonify({'error': 'Missing required parameters'}), 400

    try:
        # Retrieve file info from session
        file_info = session.get(f'insights_file_{file_id}')
        if not file_info:
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        # Handle both old format (string) and new format (dict)
        if isinstance(file_info, str):
            file_path = file_info
        else:
            file_path = file_info.get('path')

        level = int(data.get('level', 1))
        if level not in [1, 2, 3]:
            return jsonify({'error': 'Invalid level. Must be 1, 2, or 3.'}), 400

        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        # Prepare data at the same level
        result = prepare_data_for_insights(file_path, level=level)
        df_exploded = result['df_exploded']

        # Perform analysis with level for universal mean calculation
        analysis_result = analyze_imdrf_insights(
            df_exploded,
            prefix,
            manufacturers,
            grain,
            threshold_k,
            level=level
        )

        # Convert pandas data to JSON-serializable format
        manufacturer_series = {}
        for mfr, series in analysis_result['manufacturer_series'].items():
            manufacturer_series[mfr] = {
                'dates': series.index.strftime('%Y-%m-%d').tolist(),
                'values': series.tolist()
            }

        response_data = {
            'success': True,
            'universal_mean': analysis_result['universal_mean'],
            'prefix_mean': analysis_result['prefix_mean'],
            'prefix_std': analysis_result['prefix_std'],
            'upper_threshold': analysis_result['upper_threshold'],
            'lower_threshold': analysis_result['lower_threshold'],
            'manufacturer_series': manufacturer_series,
            'date_range': analysis_result['date_range'].strftime('%Y-%m-%d').tolist() if len(analysis_result['date_range']) > 0 else [],
            'statistics': analysis_result['statistics'],
            'grain': grain,
            'selected_prefix': prefix,
            'level': level,
            'level_label': analysis_result.get('level_label', f'Level-{level}')
        }

        return jsonify(response_data), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/annex-status', methods=['GET'])
@login_required
def api_annex_status():
    """Get the status of the IMDRF Annex file loading."""
    try:
        status = get_annex_status()
        return jsonify({
            'success': True,
            'annex_status': status
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/download-code-counts-xlsx', methods=['GET'])
@login_required
def api_download_code_counts_xlsx():
    """Download IMDRF code counts for all levels as an XLSX file."""
    file_id = request.args.get('file_id')

    if not file_id:
        return jsonify({'error': 'Missing file_id parameter'}), 400

    try:
        file_info = session.get(f'insights_file_{file_id}')
        if not file_info:
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        if isinstance(file_info, str):
            file_path = file_info
        else:
            file_path = file_info.get('path')

        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        counts_by_level = get_imdrf_code_counts_all_levels(file_path)

        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "IMDRF Code Counts"

        bold_font = Font(bold=True)

        for level in [1, 2, 3]:
            level_label = f"Level-{level} Codes"
            ws.append([level_label, ""])
            ws.cell(row=ws.max_row, column=1).font = bold_font

            level_counts = counts_by_level.get(level, {})
            for code in sorted(level_counts.keys()):
                ws.append([code, level_counts.get(code, 0)])

            ws.append(["", ""])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='imdrf_code_counts_all_levels.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 400


# TXT to CSV Converter Routes
@app.route('/txt-to-csv')
@login_required
def txt_to_csv_page():
    """Render TXT to CSV converter page."""
    return render_template('txt_to_csv.html', user=current_user)


@app.route('/api/txt-to-csv/upload-chunk', methods=['POST'])
@login_required
def api_txt_upload_chunk():
    """Handle chunked file upload for large TXT files (10GB+ support)."""
    try:
        chunk = request.files.get('chunk')
        chunk_index = int(request.form.get('chunkIndex', 0))
        total_chunks = int(request.form.get('totalChunks', 1))
        file_id = request.form.get('fileId')
        original_filename = request.form.get('filename', 'upload.txt')

        if not chunk:
            return jsonify({'error': 'No chunk data received'}), 400

        # Create file_id on first chunk
        if chunk_index == 0:
            file_id = f"{current_user.id}_{os.urandom(8).hex()}"

        if not file_id:
            return jsonify({'error': 'Missing file ID'}), 400

        # Secure the filename
        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")

        # Append chunk to file (create on first chunk, append on subsequent)
        mode = 'wb' if chunk_index == 0 else 'ab'
        with open(file_path, mode) as f:
            # Stream the chunk in smaller pieces to avoid memory issues
            while True:
                data = chunk.read(8192)  # 8KB at a time
                if not data:
                    break
                f.write(data)

        # Return file_id and status
        is_complete = (chunk_index + 1) >= total_chunks

        return jsonify({
            'success': True,
            'fileId': file_id,
            'chunkIndex': chunk_index,
            'isComplete': is_complete
        }), 200

    except Exception as e:
        return jsonify({'error': f'Chunk upload failed: {str(e)}'}), 400


@app.route('/api/txt-to-csv/preview', methods=['POST'])
@login_required
def api_txt_preview():
    """Get preview for already uploaded TXT file (by file_id) or upload small file."""
    try:
        # Check if this is a file_id based preview request (for chunked uploads)
        if request.is_json:
            data = request.get_json()
            file_id = data.get('file_id')
            original_filename = data.get('filename', 'upload.txt')

            if file_id:
                filename = secure_filename(original_filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")

                if not os.path.exists(file_path):
                    return jsonify({'error': 'File not found. Please upload again.'}), 404

                # Get preview
                preview = get_txt_preview(file_path, num_rows=10)

                # Store file path in session
                session[f'txt_file_{file_id}'] = file_path

                return jsonify({
                    'success': True,
                    'file_id': file_id,
                    'preview': preview
                }), 200

        # Fallback: Handle traditional file upload for smaller files
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        if not file.filename.lower().endswith('.txt'):
            return jsonify({'error': 'Invalid file type. Please upload a .txt file.'}), 400

        # Save uploaded file with unique name
        filename = secure_filename(file.filename)
        file_id = f"{current_user.id}_{os.urandom(8).hex()}"
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")

        # Stream file to disk in chunks
        with open(input_path, 'wb') as f:
            while True:
                chunk = file.read(8192)  # 8KB chunks
                if not chunk:
                    break
                f.write(chunk)

        # Get preview
        preview = get_txt_preview(input_path, num_rows=10)

        # Store file path in session
        session[f'txt_file_{file_id}'] = input_path

        return jsonify({
            'success': True,
            'file_id': file_id,
            'preview': preview
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/txt-to-csv/convert', methods=['POST'])
@login_required
def api_txt_convert():
    """Convert TXT file to CSV."""
    data = request.get_json()
    file_id = data.get('file_id')

    if not file_id:
        return jsonify({'error': 'Missing file_id parameter'}), 400

    try:
        # Retrieve file path from session
        file_path = session.get(f'txt_file_{file_id}')
        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        # Generate output path
        output_file_id = f"{current_user.id}_{os.urandom(8).hex()}"
        output_filename = os.path.splitext(os.path.basename(file_path))[0] + '.csv'
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{output_file_id}_{output_filename}")

        # Convert file
        converter = TxtToCsvConverter()
        stats = converter.process_file_chunked(file_path, output_path)

        # Store output path in session
        session[f'csv_file_{output_file_id}'] = {
            'path': output_path,
            'filename': output_filename
        }

        return jsonify({
            'success': True,
            'output_file_id': output_file_id,
            'stats': stats
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/txt-to-csv/download/<file_id>')
@login_required
def api_txt_download(file_id):
    """Download converted CSV file."""
    try:
        # Retrieve file info from session
        file_info = session.get(f'csv_file_{file_id}')
        if not file_info or not os.path.exists(file_info['path']):
            return jsonify({'error': 'File not found. Please convert again.'}), 404

        return send_file(
            file_info['path'],
            as_attachment=True,
            download_name=file_info['filename'],
            mimetype='text/csv'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 400


# CSV Viewer Routes
@app.route('/csv-viewer')
@login_required
def csv_viewer_page():
    """Render CSV viewer page."""
    return render_template('csv_viewer.html', user=current_user)


@app.route('/api/csv-viewer/upload-chunk', methods=['POST'])
@login_required
def api_csv_upload_chunk():
    """Handle chunked file upload for CSV viewer."""
    try:
        chunk = request.files.get('chunk')
        chunk_index = int(request.form.get('chunkIndex', 0))
        total_chunks = int(request.form.get('totalChunks', 1))
        file_id = request.form.get('fileId')
        original_filename = request.form.get('filename', 'upload.csv')

        if not chunk:
            return jsonify({'error': 'No chunk data received'}), 400

        # Create file_id on first chunk
        if chunk_index == 0:
            file_id = f"{current_user.id}_{os.urandom(8).hex()}"

        if not file_id:
            return jsonify({'error': 'Missing file ID'}), 400

        # Secure the filename
        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        # Append chunk to file
        mode = 'wb' if chunk_index == 0 else 'ab'
        with open(file_path, mode) as f:
            while True:
                data = chunk.read(8192)
                if not data:
                    break
                f.write(data)

        return jsonify({
            'success': True,
            'fileId': file_id,
            'chunkIndex': chunk_index,
            'isComplete': (chunk_index + 1) >= total_chunks
        }), 200

    except Exception as e:
        return jsonify({'error': f'Chunk upload failed: {str(e)}'}), 400


@app.route('/api/csv-viewer/info', methods=['POST'])
@login_required
def api_csv_info():
    """Get CSV file information."""
    try:
        data = request.get_json()
        file_id = data.get('file_id')
        original_filename = data.get('filename', 'upload.csv')

        if not file_id:
            return jsonify({'error': 'Missing file_id'}), 400

        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404

        # Store in session
        session[f'csv_viewer_{file_id}'] = file_path

        # Get file info
        info = get_csv_info(file_path)

        return jsonify({
            'success': True,
            'info': info
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/csv-viewer/page', methods=['POST'])
@login_required
def api_csv_page():
    """Get a page of CSV data."""
    try:
        data = request.get_json()
        file_id = data.get('file_id')
        original_filename = data.get('filename', 'upload.csv')
        page = data.get('page', 1)
        page_size = data.get('page_size', 100)

        if not file_id:
            return jsonify({'error': 'Missing file_id'}), 400

        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404

        # Get page data
        page_data = get_csv_page(file_path, page, page_size)

        return jsonify(page_data), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/csv-viewer/search', methods=['POST'])
@login_required
def api_csv_search():
    """Search in CSV file."""
    try:
        data = request.get_json()
        file_id = data.get('file_id')
        original_filename = data.get('filename', 'upload.csv')
        search_term = data.get('search_term', '')

        if not file_id or not search_term:
            return jsonify({'error': 'Missing file_id or search_term'}), 400

        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404

        # Search in file
        viewer = LargeCSVViewer()
        results = viewer.search_in_file(file_path, search_term)

        return jsonify(results), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/csv-viewer/column-stats', methods=['POST'])
@login_required
def api_csv_column_stats():
    """Get column statistics."""
    try:
        data = request.get_json()
        file_id = data.get('file_id')
        original_filename = data.get('filename', 'upload.csv')
        column_index = data.get('column_index', 0)

        if not file_id:
            return jsonify({'error': 'Missing file_id'}), 400

        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404

        # Get column stats
        viewer = LargeCSVViewer()
        stats = viewer.get_column_stats(file_path, column_index)

        return jsonify({
            'success': True,
            'stats': stats
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


if __name__ == '__main__':
    # Check for Groq API key
    if not GROQ_API_KEY:
        print("WARNING: GROQ_API_KEY not set. Set it as an environment variable.")
    
    # Check for mail configuration
    if not MAIL_USERNAME or not MAIL_PASSWORD:
        print("WARNING: Mail credentials not configured. Password recovery emails will not work.")
        print("Set MAIL_USERNAME and MAIL_PASSWORD environment variables.")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
